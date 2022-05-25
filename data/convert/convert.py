from datetime import datetime
from tabnanny import check
from typing import List, Dict, Any
import re
import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

months = ["January", "February", "March", "April", "May",
          "June", "July", "August", "October", "November", "December"]
year_regex = r'^\d{4}$'
month_regex = re.compile('^(' + '|'.join(months) + "), (\\d{4})$")
day_regex = re.compile('^(' + '|'.join(months) + ") (\\d{1,2}), (\\d{4})$")

# support year-month-day and day-month-year
numeric_date_regex = r'(\d{1,4})[-/](\d{1,2})[-/](\d{1,4})'

def convert_file(in_path, out_path):
    data = parse_input(in_path)

    with open(out_path, 'w') as outfile:
        json.dump(data, outfile, indent=4, separators=(", ", ": "))


def parse_input(in_path):
    wb = load_workbook(in_path)

    data = {
        sheetname.lower(): parse_sheet(wb[sheetname])
        for sheetname in wb.sheetnames
    }

    check_relations(data)
    check_dates(data)
    return data


def check_relations(data):
    print('\nCHECKING RELATIONS...')
    print("Checking authors:")
    for author in data['authors']:
        check_dependency(data, 'locations', author['place_of_birth'])
        check_dependency(data, 'locations', author['place_of_death'])

    print("Checking works:")
    for work in data['works']:
        check_dependency(data, 'authors', work['author_name'])
        check_dependency(data, 'locations', work['where'])
        check_required(work, 'author_name')

    print("Checking legacy:")
    for legacy in data['legacy']:
        check_dependency(data, 'authors', legacy['about'])
        check_dependency(data, 'locations', legacy['where'])
        check_required(legacy, 'about')

    print("Checking events:")
    for event in data['events']:
        check_dependency(data, 'authors', event['author'])
        check_dependency(data, 'locations', event['where'])
        check_required(event, 'author')


def check_dependency(data, sheet, name):
    if name is None:
        return
    for item in data[sheet]:
        if name == item['name']:
            return

    print(f"! {name} does not exist in {sheet}")

def check_required(item, name):
    if not item[name]:
        print(f'{name} missing for {item}')

def check_dates(data):
    "For events with start + end date, check that end > start"
    print('\nCHECKING DATES...')
    print('Checking works:')

    for work in data['works']:
        check_start_end_date(work)
    for legacy in data['legacy']:
        check_start_end_date(legacy)
    for event in data['events']:
        check_start_end_date(event)

def check_start_end_date(item):
    if 'start_date' in item and item['start_date'] and 'end_date' in item and item['end_date']:
        start_date = format_date(item['start_date'])
        end_date = format_date(item['end_date'])
        start_year = int(start_date[:4])
        end_year = int(end_date [:4])

        if end_year < start_year:
            title = item['title']
            print(f'! End year {end_year} precedes start year {start_year} for "{title}"')
    return

def parse_sheet(worksheet: Worksheet):
    header_row = next(worksheet.rows)

    def format_header(header: str):
        def lowercase(h): return h.lower()
        def strip(h): return h.strip()
        def underscore(h): return '_'.join(h.split())
        return underscore(strip(lowercase(header)))

    def format_row(row):
        values = [cell.value for cell in row]
        return {
            header: format_value(value, header) for header, value in zip(headers, values)
        }

    headers = [format_header(cell.value)
               for cell in filter(lambda cell: cell.value is not None,
                                  header_row)]

    data = [format_row(row) for row in list(worksheet.rows)[1:]]

    return list(skip_empty(data))


def skip_empty(data: List[Dict[str, Any]]):
    for item in data:
        for value in item.values():
            if value is not None:
                yield item
                break


def format_date(value):
    """"
    Formats date as either:

    * year e.g. 2022
    * year-month e.g. 2022-3
    * year-month-day e.g. 2022-3-30
    """
    date = None

    if type(value) == int:
        return f"{value}"
    elif type(value) == datetime:
        date = value
    elif type(value) == str:
        value = value.strip()
        if not value:
            return None
        if value == 'Toevoegen': # used as temporary filler
            return None

        match = re.search(numeric_date_regex, value)
        if match:
            parts = match.groups()
            if int(parts[0]) > 31:
                year, month, day = parts
            else:
                day, month, year = parts
            date = datetime(
                year=int(year),
                month=int(month),
                day=int(day)
            )
        else:
            if re.search(year_regex, value):
                return value

            match = month_regex.search(value)
            if match:
                month, year = match.groups()
                return f"{year}-{months.index(month) + 1}"

            match = day_regex.search(value)
            if match:
                month, day, year = match.groups()
                date = datetime(
                    year=int(year),
                    month=int(months.index(month) + 1),
                    day=int(day))

    if date is not None:
        return date.strftime('%Y-%m-%d')

    raise Exception("Unsupported type " + str(type(value)))


def format_value(value, header):
    if value is None:
        return None

    if header in ['lat', 'long']:
        if type(value) == float:
            # floats are used for for lat/long coordinates
            return value
        if type(value) == str:
            stripped = value.strip('() ') # remove whitespace and parentheses
            return float(stripped)

    if header in ['date', 'date_of_birth', 'date_of_death', 'start_date', 'end_date']:
        return format_date(value)

    if type(value) == str:
        # remove leading/trailing whitespace for string values
        return value.strip()

    if type(value) == int:
        return value

    return ''
